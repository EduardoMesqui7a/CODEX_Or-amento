import io
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from rapidfuzz import fuzz
from sentence_transformers import SentenceTransformer
from sklearn.neighbors import NearestNeighbors
from unidecode import unidecode

st.set_page_config(page_title="Orçamento IA - VSN", layout="wide")

MODELO_EMBEDDING = "sentence-transformers/all-MiniLM-L6-v2"
PESO_SEMANTICO = 0.70
PESO_FUZZY = 0.20
PESO_REGRAS = 0.10


@st.cache_resource
def carregar_modelo():
    return SentenceTransformer(MODELO_EMBEDDING)


def normalizar_texto(texto: str) -> str:
    if texto is None or pd.isna(texto):
        return ""

    texto = str(texto).strip().lower()
    texto = unidecode(texto)

    substituicoes = {
        "fck": "resistencia caracteristica",
        "mpa": "megapascal",
        "concreto armado": "concreto estrutural armado",
        "concreto simples": "concreto sem armadura",
        "divisoria": "parede divisoria vedacao compartimentacao interna",
        "drywall": "parede leve em gesso acartonado",
        "alvenaria": "parede de alvenaria vedacao",
        "parede": "vedacao parede fechamento",
        "aco": "aco armadura",
        "armacao": "armadura aco",
        "forma": "forma madeira compensado",
        "tubo": "tubulacao",
        "tubos": "tubulacao",
        "eletroduto": "tubulacao eletrica conduite",
        "conduite": "tubulacao eletrica eletroduto",
        "piso": "pavimentacao revestimento piso",
        "bloco": "alvenaria bloco",
        "reboco": "argamassa revestimento",
        "chapisco": "argamassa aderencia",
        "escavacao": "movimento de terra escavacao",
        "aterro": "movimento de terra aterro compactacao",
        "lastro": "camada de regularizacao lastro",
    }

    for de, para in substituicoes.items():
        texto = texto.replace(de, para)

    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def score_regras(busca_norm: str, descricao_norm: str) -> float:
    score = 0.0

    numeros_relevantes = ["5", "8", "10", "12", "15", "20", "25", "30", "35", "40", "50"]
    for numero in numeros_relevantes:
        if numero in busca_norm and numero in descricao_norm:
            score += 0.10

    pares = [
        ("concreto", "concreto"),
        ("armado", "armado"),
        ("argamassa", "argamassa"),
        ("alvenaria", "alvenaria"),
        ("divisoria", "divisoria"),
        ("drywall", "drywall"),
        ("piso", "piso"),
        ("tubulacao", "tubulacao"),
        ("eletrica", "eletrica"),
        ("hidraulica", "hidraulica"),
        ("escavacao", "escavacao"),
        ("aterro", "aterro"),
        ("forma", "forma"),
        ("aco", "aco"),
        ("vedacao", "vedacao"),
        ("bloco", "bloco"),
        ("porta", "porta"),
        ("janela", "janela"),
    ]

    for termo_busca, termo_desc in pares:
        if termo_busca in busca_norm and termo_desc in descricao_norm:
            score += 0.08

    if "divisoria" in busca_norm:
        if any(x in descricao_norm for x in ["drywall", "alvenaria", "parede", "vedacao"]):
            score += 0.20

    if "concreto" in busca_norm and "megapascal" in busca_norm:
        if "concreto" in descricao_norm and any(
            x in descricao_norm for x in ["megapascal", "resistencia caracteristica"]
        ):
            score += 0.20

    return min(score, 1.0)


def eh_linha_de_titulo_ou_subtitulo(texto) -> bool:
    if texto is None or str(texto).strip() == "":
        return True

    t = str(texto).strip()
    t_norm = normalizar_texto(t)
    palavras = t_norm.split()

    if len(t_norm) <= 3:
        return True

    termos_genericos = {
        "servicos preliminares",
        "fundacoes",
        "estrutura",
        "superestrutura",
        "arquitetura",
        "instalacoes",
        "instalacoes eletricas",
        "instalacoes hidraulicas",
        "urbanizacao",
        "cobertura",
        "revestimentos",
        "esquadrias",
        "pintura",
        "demolicoes",
        "demolicao",
        "movimento de terra",
        "infraestrutura",
        "equipamentos",
        "geral",
        "administracao local",
    }
    if t_norm in termos_genericos:
        return True

    unidades = {"m", "m2", "m3", "kg", "un", "vb", "cj", "h", "mes"}
    tem_numero = bool(re.search(r"\d", t_norm))
    tem_unidade = any(u in palavras for u in unidades)

    if len(palavras) <= 3 and not tem_numero and not tem_unidade:
        return True

    letras = [c for c in t if c.isalpha()]
    if letras:
        proporcao_maiuscula = sum(1 for c in letras if c.isupper()) / len(letras)
        if proporcao_maiuscula > 0.8 and len(palavras) <= 5:
            return True

    return False


def carregar_excel(uploaded_file, nome_aba: Optional[str], header_index: int) -> pd.DataFrame:
    uploaded_file.seek(0)
    xls = pd.ExcelFile(uploaded_file)
    aba = nome_aba if nome_aba else xls.sheet_names[0]

    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=aba, header=header_index)
    df.columns = [str(c).strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False)
def preparar_base_para_busca(df_base: pd.DataFrame, coluna_texto_base: str):
    modelo = carregar_modelo()

    df_base_proc = df_base.copy()
    df_base_proc[coluna_texto_base] = df_base_proc[coluna_texto_base].fillna("").astype(str)
    df_base_proc["__texto_base_norm__"] = df_base_proc[coluna_texto_base].map(normalizar_texto)

    textos_norm = df_base_proc["__texto_base_norm__"].tolist()

    embeddings = modelo.encode(
        textos_norm,
        convert_to_numpy=True,
        normalize_embeddings=True,
        show_progress_bar=False,
        batch_size=128,
    )

    indice = NearestNeighbors(metric="cosine", algorithm="auto")
    indice.fit(embeddings)

    return df_base_proc, embeddings, indice


def buscar_melhor_item_em_lote(
    buscas_norm_unicas: List[str],
    df_base_proc: pd.DataFrame,
    indice,
    top_k_candidatos: int,
) -> Dict[str, Optional[Tuple[int, dict]]]:
    modelo = carregar_modelo()

    if not buscas_norm_unicas:
        return {}

    emb_buscas = modelo.encode(
        buscas_norm_unicas,
        convert_to_numpy=True,
        normalize_embeddings=True,
        show_progress_bar=False,
        batch_size=128,
    )

    k = min(top_k_candidatos, len(df_base_proc))
    distancias_lote, indices_lote = indice.kneighbors(emb_buscas, n_neighbors=k)

    resultados = {}

    for pos_busca, busca_norm in enumerate(buscas_norm_unicas):
        melhores_indices = indices_lote[pos_busca]
        melhores_distancias = distancias_lote[pos_busca]

        melhor_idx = None
        melhor_score = -1.0
        melhor_det = None

        for pos_cand, idx_base in enumerate(melhores_indices):
            texto_base_norm = df_base_proc.iloc[idx_base]["__texto_base_norm__"]

            score_sem = 1.0 - float(melhores_distancias[pos_cand])
            score_fuzzy = fuzz.token_set_ratio(busca_norm, texto_base_norm) / 100.0
            score_reg = score_regras(busca_norm, texto_base_norm)

            score_final = (
                PESO_SEMANTICO * score_sem
                + PESO_FUZZY * score_fuzzy
                + PESO_REGRAS * score_reg
            )

            if score_final > melhor_score:
                melhor_score = score_final
                melhor_idx = int(idx_base)
                melhor_det = {
                    "score_final": round(score_final, 4),
                    "score_semantico": round(score_sem, 4),
                    "score_fuzzy": round(score_fuzzy, 4),
                    "score_regras": round(score_reg, 4),
                }

        if melhor_idx is None:
            resultados[busca_norm] = None
        else:
            resultados[busca_norm] = (melhor_idx, melhor_det)

    return resultados


def obter_celula_segura_para_escrita(ws, linha: int, coluna: int):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= linha <= merged_range.max_row
            and merged_range.min_col <= coluna <= merged_range.max_col
        ):
            if merged_range.min_row == merged_range.max_row == linha:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return None

    return ws.cell(row=linha, column=coluna)


def encontrar_ultima_coluna_com_dados(ws) -> int:
    for coluna in range(ws.max_column, 0, -1):
        for linha in range(1, ws.max_row + 1):
            valor = ws.cell(row=linha, column=coluna).value
            if valor is not None and str(valor).strip() != "":
                return coluna
    return 0


def obter_nome_coluna_referencia(coluna_texto_base: str) -> str:
    return f"IA_REFERENCIA_BASE ({coluna_texto_base})"


def aplicar_resultado_no_excel_original(
    uploaded_file,
    nome_aba: str,
    header_index: int,
    df_original: pd.DataFrame,
    df_resultado: pd.DataFrame,
    colunas_destino_preencher: List[str],
    nome_coluna_referencia: str,
) -> bytes:
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb[nome_aba] if nome_aba in wb.sheetnames else wb[wb.sheetnames[0]]

    linha_cabecalho_excel = header_index + 1
    primeira_linha_dados_excel = linha_cabecalho_excel + 1

    mapa_colunas_destino = {}
    for nome_coluna in colunas_destino_preencher:
        indice_df = df_original.columns.get_loc(nome_coluna) + 1
        mapa_colunas_destino[nome_coluna] = indice_df

    ultima_coluna_com_dados = encontrar_ultima_coluna_com_dados(ws)
    coluna_referencia_excel = ultima_coluna_com_dados + 1
    ws.cell(row=linha_cabecalho_excel, column=coluna_referencia_excel).value = nome_coluna_referencia

    for i in range(len(df_resultado)):
        linha_excel = primeira_linha_dados_excel + i

        for nome_coluna in colunas_destino_preencher:
            if nome_coluna not in df_resultado.columns:
                continue

            coluna_excel = mapa_colunas_destino[nome_coluna]
            valor = df_resultado.iloc[i][nome_coluna]

            celula_destino = obter_celula_segura_para_escrita(ws, linha_excel, coluna_excel)
            if celula_destino is None:
                continue

            celula_destino.value = valor

        if nome_coluna_referencia in df_resultado.columns:
            celula_referencia = obter_celula_segura_para_escrita(ws, linha_excel, coluna_referencia_excel)
            if celula_referencia is not None:
                celula_referencia.value = df_resultado.iloc[i][nome_coluna_referencia]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def processar_preenchimento(
    df_base: pd.DataFrame,
    df_destino: pd.DataFrame,
    coluna_busca_destino: str,
    colunas_base_retorno: List[str],
    colunas_destino_preencher: List[str],
    coluna_texto_base: str,
    score_minimo: float,
    top_k_candidatos: int,
):
    df_destino_proc = df_destino.copy()

    score_col = "IA_SCORE"
    match_col = "IA_DESCRICAO_ENCONTRADA"
    idx_col = "IA_LINHA_BASE"
    tipo_col = "IA_TIPO_LINHA"
    referencia_col = obter_nome_coluna_referencia(coluna_texto_base)

    for col in [score_col, match_col, idx_col, tipo_col, referencia_col]:
        if col not in df_destino_proc.columns:
            df_destino_proc[col] = None

    status = st.empty()
    progresso = st.progress(0)

    status.info("Em processamento. Analisando a base de dados e preparando a busca semântica.")
    df_base_proc, embeddings, indice = preparar_base_para_busca(df_base, coluna_texto_base)
    progresso.progress(0.10)

    if coluna_busca_destino not in df_destino_proc.columns:
        status.warning("A coluna de busca não foi encontrada na planilha de destino.")
        progresso.empty()
        return df_destino_proc

    total = len(df_destino_proc)
    buscas_originais = df_destino_proc[coluna_busca_destino].tolist()

    status.info("Em processamento. Analisando a planilha de destino e separando as buscas válidas.")
    mapa_buscas_validas: Dict[str, str] = {}
    for busca in buscas_originais:
        if busca is None or str(busca).strip() == "":
            continue
        if eh_linha_de_titulo_ou_subtitulo(busca):
            continue

        busca_str = str(busca)
        busca_norm = normalizar_texto(busca_str)
        if busca_norm:
            mapa_buscas_validas[busca_str] = busca_norm

    buscas_norm_unicas = list(set(mapa_buscas_validas.values()))
    progresso.progress(0.25)

    status.info("Em processamento. Calculando as melhores correspondências da base de dados.")
    resultados_unicos = buscar_melhor_item_em_lote(
        buscas_norm_unicas=buscas_norm_unicas,
        df_base_proc=df_base_proc,
        indice=indice,
        top_k_candidatos=top_k_candidatos,
    )
    progresso.progress(0.55)

    cache_busca_original: Dict[str, Optional[Tuple[int, dict]]] = {}
    for busca_original, busca_norm in mapa_buscas_validas.items():
        cache_busca_original[busca_original] = resultados_unicos.get(busca_norm)

    status.info("Em processamento. Preenchendo os dados na planilha de destino.")

    for i, busca in enumerate(buscas_originais):
        if busca is None or str(busca).strip() == "":
            df_destino_proc.at[i, tipo_col] = "Vazia"
            if i % 25 == 0 or i == total - 1:
                progresso.progress(0.55 + 0.45 * ((i + 1) / max(total, 1)))
            continue

        if eh_linha_de_titulo_ou_subtitulo(busca):
            df_destino_proc.at[i, tipo_col] = "Título/Subtítulo"
            if i % 25 == 0 or i == total - 1:
                progresso.progress(0.55 + 0.45 * ((i + 1) / max(total, 1)))
            continue

        res = cache_busca_original.get(str(busca))

        if res is None:
            df_destino_proc.at[i, tipo_col] = "Sem correspondência"
            df_destino_proc.at[i, referencia_col] = "Sem correspondência"
            if i % 25 == 0 or i == total - 1:
                progresso.progress(0.55 + 0.45 * ((i + 1) / max(total, 1)))
            continue

        idx_match, det = res
        referencia_base = df_base_proc.iloc[idx_match][coluna_texto_base]

        if det["score_final"] < score_minimo:
            df_destino_proc.at[i, score_col] = det["score_final"]
            df_destino_proc.at[i, match_col] = "Confiança baixa"
            df_destino_proc.at[i, idx_col] = int(idx_match) + 2
            df_destino_proc.at[i, tipo_col] = "Item, confiança baixa"
            df_destino_proc.at[i, referencia_col] = referencia_base
            if i % 25 == 0 or i == total - 1:
                progresso.progress(0.55 + 0.45 * ((i + 1) / max(total, 1)))
            continue

        for col_base, col_dest in zip(colunas_base_retorno, colunas_destino_preencher):
            df_destino_proc.at[i, col_dest] = df_base_proc.iloc[idx_match][col_base]

        df_destino_proc.at[i, score_col] = det["score_final"]
        df_destino_proc.at[i, match_col] = referencia_base
        df_destino_proc.at[i, idx_col] = int(idx_match) + 2
        df_destino_proc.at[i, tipo_col] = "Item"
        df_destino_proc.at[i, referencia_col] = referencia_base

        if i % 25 == 0 or i == total - 1:
            status.info(f"Em processamento. Preenchendo os dados na planilha de destino. Linha {i + 1} de {total}.")
            progresso.progress(0.55 + 0.45 * ((i + 1) / max(total, 1)))

    progresso.progress(1.0)
    status.success("Processamento concluído. Planilha analisada e dados preenchidos.")
    return df_destino_proc


st.title("Orçamento IA - VSN")
st.caption("Importe a base de dados e a planilha a preencher, escolha as colunas e gere o arquivo preenchido.")

with st.sidebar:
    st.header("Configurações")
    score_minimo = st.slider("Score mínimo para preencher", 0.0, 1.0, 0.35, 0.01)
    header_base = st.number_input("Linha do cabeçalho da base", min_value=1, value=1, step=1)
    header_dest = st.number_input("Linha do cabeçalho da planilha a preencher", min_value=1, value=1, step=1)
    st.markdown("Sugestão, se a base tem cabeçalho na linha 3 do Excel, informe 3.")
    top_k_candidatos = st.number_input("Qtd. de candidatos por busca", min_value=5, max_value=100, value=30, step=5)

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Base de dados")
    arquivo_base = st.file_uploader("Importar base de dados", type=["xlsx", "xlsm", "xls"], key="base")

with col2:
    st.subheader("2. Planilha a preencher")
    arquivo_destino = st.file_uploader("Importar planilha de destino", type=["xlsx", "xlsm", "xls"], key="destino")

if arquivo_base and arquivo_destino:
    try:
        arquivo_base.seek(0)
        xls_base = pd.ExcelFile(arquivo_base)

        arquivo_destino.seek(0)
        xls_dest = pd.ExcelFile(arquivo_destino)

        col3, col4 = st.columns(2)
        with col3:
            aba_base = st.selectbox("Aba da base", options=xls_base.sheet_names)
        with col4:
            aba_dest = st.selectbox("Aba da planilha a preencher", options=xls_dest.sheet_names)

        df_base = carregar_excel(arquivo_base, aba_base, int(header_base) - 1)
        df_destino = carregar_excel(arquivo_destino, aba_dest, int(header_dest) - 1)

        st.divider()
        st.subheader("3. Mapeamento das colunas")

        c1, c2 = st.columns(2)
        with c1:
            coluna_texto_base = st.selectbox(
                "Coluna da base usada para comparação semântica",
                options=df_base.columns.tolist(),
                index=df_base.columns.tolist().index("DESCRIÇÃO") if "DESCRIÇÃO" in df_base.columns else 0,
            )
        with c2:
            coluna_busca_destino = st.selectbox(
                "Coluna da planilha de destino usada como busca",
                options=df_destino.columns.tolist(),
                index=df_destino.columns.tolist().index("G") if "G" in df_destino.columns else 0,
            )

        st.info(
            f"A referência do item encontrado será gravada em uma nova coluna no final da planilha, "
            f"usando os valores da coluna '{coluna_texto_base}' da base."
        )

        st.markdown("### Colunas da base que deseja obter")
        colunas_base_retorno = st.multiselect(
            "Selecione as colunas da base",
            options=df_base.columns.tolist(),
            default=[c for c in ["R$ CAPEX/NOVO", "CÓDIGO", "FONTE", "UNID", "SEM BDI"] if c in df_base.columns],
        )

        st.markdown("### Colunas da planilha de destino que receberão os dados")
        st.caption("A ordem deve corresponder exatamente à ordem escolhida na base.")

        colunas_destino_preencher = []
        for i, col_base in enumerate(colunas_base_retorno, start=1):
            escolha = st.selectbox(
                f"Destino para '{col_base}'",
                options=df_destino.columns.tolist(),
                key=f"dest_{i}_{col_base}",
            )
            colunas_destino_preencher.append(escolha)

        if len(colunas_base_retorno) != len(colunas_destino_preencher):
            st.error("A quantidade de colunas da base e de destino precisa ser a mesma.")
        elif len(set(colunas_destino_preencher)) != len(colunas_destino_preencher):
            st.error("Você repetiu colunas de destino. Cada coluna de destino deve ser usada apenas uma vez.")
        else:
            st.divider()
            st.subheader("4. Prévia")

            p1, p2 = st.columns(2)
            with p1:
                st.write("Base")
                st.dataframe(df_base.head(10), use_container_width=True)
            with p2:
                st.write("Planilha a preencher")
                st.dataframe(df_destino.head(10), use_container_width=True)

            if st.button("Processar preenchimento", type="primary"):
                resultado = processar_preenchimento(
                    df_base=df_base,
                    df_destino=df_destino,
                    coluna_busca_destino=coluna_busca_destino,
                    colunas_base_retorno=colunas_base_retorno,
                    colunas_destino_preencher=colunas_destino_preencher,
                    coluna_texto_base=coluna_texto_base,
                    score_minimo=score_minimo,
                    top_k_candidatos=int(top_k_candidatos),
                )

                st.success("Processamento concluído.")
                st.dataframe(resultado.head(50), use_container_width=True)

                excel_bytes = aplicar_resultado_no_excel_original(
                    uploaded_file=arquivo_destino,
                    nome_aba=aba_dest,
                    header_index=int(header_dest) - 1,
                    df_original=df_destino,
                    df_resultado=resultado,
                    colunas_destino_preencher=colunas_destino_preencher,
                    nome_coluna_referencia=obter_nome_coluna_referencia(coluna_texto_base),
                )

                st.download_button(
                    label="Baixar planilha preenchida",
                    data=excel_bytes,
                    file_name="planilha_preenchida.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    except Exception as e:
        st.error(f"Erro ao processar os arquivos: {e}")
else:
    st.info("Importe os dois arquivos para habilitar o mapeamento e o preenchimento automático.")
