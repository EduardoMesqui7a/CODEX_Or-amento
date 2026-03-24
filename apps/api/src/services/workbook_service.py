from typing import Any, Dict, List

from openpyxl import load_workbook

from .file_service import FileService


class WorkbookInspectService:
    def __init__(self, file_service: FileService):
        self.file_service = file_service

    def list_sheet_names(self, file_id: str, user_id: str):
        db_file = self.file_service.get_file(file_id, user_id)
        stream = self.file_service.open_as_bytesio(db_file)
        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            return {"file_id": file_id, "sheets": workbook.sheetnames}
        finally:
            workbook.close()

    @staticmethod
    def _sheet_preview(ws, header_row: int, preview_limit: int = 3) -> Dict[str, Any]:
        header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
        columns = [str(col).strip() if col is not None else "" for col in header_values]
        preview_rows: List[Dict[str, Any]] = []

        for row_values in ws.iter_rows(
            min_row=header_row + 1,
            max_row=header_row + preview_limit,
            values_only=True,
        ):
            row = {}
            for idx, col in enumerate(columns):
                if not col:
                    continue
                value = row_values[idx] if idx < len(row_values) else None
                row[col] = value if value is not None else ""
            preview_rows.append(row)

        return {"columns": columns, "preview_rows": preview_rows}

    def inspect(self, file_id: str, user_id: str, header_row: int = 1):
        db_file = self.file_service.get_file(file_id, user_id)
        stream = self.file_service.open_as_bytesio(db_file)
        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            previews = []
            sheet_names = workbook.sheetnames
            for sheet_name in sheet_names[:1]:
                try:
                    ws = workbook[sheet_name]
                    preview = self._sheet_preview(ws, header_row, preview_limit=3)
                    previews.append(
                        {
                            "sheet_name": sheet_name,
                            "columns": preview["columns"],
                            "preview_rows": preview["preview_rows"],
                        }
                    )
                except Exception:
                    continue

            return {"file_id": file_id, "sheets": sheet_names, "previews": previews}
        finally:
            workbook.close()

    def inspect_sheet(self, file_id: str, user_id: str, sheet_name: str, header_row: int = 1):
        db_file = self.file_service.get_file(file_id, user_id)
        stream = self.file_service.open_as_bytesio(db_file)
        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            preview = self._sheet_preview(ws, header_row, preview_limit=5)
            return {"sheet_name": sheet_name, "columns": preview["columns"], "preview_rows": preview["preview_rows"]}
        finally:
            workbook.close()
