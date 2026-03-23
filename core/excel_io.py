import io
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook


def carregar_excel(uploaded_file, nome_aba: Optional[str], header_index: int) -> pd.DataFrame:
    uploaded_file.seek(0)
    xls = pd.ExcelFile(uploaded_file)
    aba = nome_aba if nome_aba else xls.sheet_names[0]
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=aba, header=header_index)
    df.columns = [str(coluna).strip() for coluna in df.columns]
    return df


def obter_celula_segura_para_escrita(ws, linha: int, coluna: int):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= linha <= merged_range.max_row
            and merged_range.min_col <= coluna <= merged_range.max_col
        ):
            if merged_range.min_row == merged_range.max_row == linha:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return None
    return ws.cell(row=linha, column=coluna)


def encontrar_ultima_coluna_com_dados(ws) -> int:
    for coluna in range(ws.max_column, 0, -1):
        for linha in range(1, ws.max_row + 1):
            valor = ws.cell(row=linha, column=coluna).value
            if valor is not None and str(valor).strip() != "":
                return coluna
    return 0


def obter_nome_coluna_referencia(coluna_texto_base: str) -> str:
    return f"IA_REFERENCIA_BASE ({coluna_texto_base})"


def aplicar_resultado_no_excel_original(
    uploaded_file,
    nome_aba: str,
    header_index: int,
    df_original: pd.DataFrame,
    df_resultado: pd.DataFrame,
    colunas_destino_preencher: List[str],
    nome_coluna_referencia: str,
) -> bytes:
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb[nome_aba] if nome_aba in wb.sheetnames else wb[wb.sheetnames[0]]

    linha_cabecalho_excel = header_index + 1
    primeira_linha_dados_excel = linha_cabecalho_excel + 1

    mapa_colunas_destino = {}
    for nome_coluna in colunas_destino_preencher:
        indice_df = df_original.columns.get_loc(nome_coluna) + 1
        mapa_colunas_destino[nome_coluna] = indice_df

    ultima_coluna_com_dados = encontrar_ultima_coluna_com_dados(ws)
    coluna_referencia_excel = ultima_coluna_com_dados + 1
    ws.cell(row=linha_cabecalho_excel, column=coluna_referencia_excel).value = nome_coluna_referencia

    for i in range(len(df_resultado)):
        linha_excel = primeira_linha_dados_excel + i
        for nome_coluna in colunas_destino_preencher:
            if nome_coluna not in df_resultado.columns:
                continue
            coluna_excel = mapa_colunas_destino[nome_coluna]
            valor = df_resultado.iloc[i][nome_coluna]
            celula_destino = obter_celula_segura_para_escrita(ws, linha_excel, coluna_excel)
            if celula_destino is not None:
                celula_destino.value = valor

        if nome_coluna_referencia in df_resultado.columns:
            celula_ref = obter_celula_segura_para_escrita(ws, linha_excel, coluna_referencia_excel)
            if celula_ref is not None:
                celula_ref.value = df_resultado.iloc[i][nome_coluna_referencia]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

